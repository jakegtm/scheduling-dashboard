# ============================================================
# processors/variance.py — OpenAir vs Schedule Variance
# ============================================================
# OpenAir file structure (verified from sample):
#   Row 1:  "Project Owner/Code" | "Week starting M/D[/YY]" x N | "Total"
#   Row 2+: Hierarchical data, differentiated by leading spaces in col A:
#             0 spaces  → Person      (e.g. "Brooks V")
#             1-2 spaces → Client     (e.g. " Vertex")   — aggregates, SKIPPED
#             4+ spaces  → Project    (e.g. "    Aramark 1099 Reporting Services")
#
# Person matching  : first token of OpenAir name → schedule column header
#                    "Brooks V" → "Brooks"
# Project matching : strip + case-insensitive, with partial-match fallback
#
# Proration logic (per user spec):
#   If a week straddles two periods (e.g. week starting May 14):
#     - Count Mon-Fri business days in each period
#     - Split hours proportionally (2 days period-1, 3 days period-2 → 2/5 + 3/5)
#
# Variance flag: |actual - scheduled| > VARIANCE_THRESHOLD hours
# ============================================================

import re
from datetime import date, timedelta
from collections import defaultdict

VARIANCE_THRESHOLD = 3   # hours


# ================================================================
# DATE / PERIOD HELPERS
# ================================================================

_WEEK_RE = re.compile(r'week\s+starting\s+(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?',
                      re.IGNORECASE)


def _parse_week_date(header: str, default_year: int) -> date | None:
    """Parse 'Week starting M/D' or 'Week starting M/D/YYYY' → date."""
    m = _WEEK_RE.search(str(header))
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    year = int(m.group(3)) if m.group(3) else default_year
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _business_days_overlap(week_start: date, p_start: date, p_end: date) -> int:
    """
    Count Mon–Fri days in the intersection of
    [week_start, week_start+6] and [p_start, p_end].
    """
    week_end   = week_start + timedelta(days=6)   # full 7-day calendar week
    overlap_s  = max(week_start, p_start)
    overlap_e  = min(week_end,   p_end)
    if overlap_s > overlap_e:
        return 0
    count = 0
    cur = overlap_s
    while cur <= overlap_e:
        if cur.weekday() < 5:    # Mon=0 … Fri=4
            count += 1
        cur += timedelta(days=1)
    return count


def _period_bounds(label: str, year: int) -> tuple[date, date] | None:
    """
    Convert a period label like 'May 1 - 15' or 'May 16-31'
    to (start_date, end_date).
    """
    import calendar
    m = re.search(
        r'([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*(\d{1,2})',
        str(label), re.IGNORECASE)
    if not m:
        return None
    _MONTHS = {mon[:3].lower(): i for i, mon in enumerate(
        ['January','February','March','April','May','June',
         'July','August','September','October','November','December'], 1)}
    mon_num = _MONTHS.get(m.group(1).lower()[:3])
    if not mon_num:
        return None
    start_day, end_day = int(m.group(2)), int(m.group(3))
    try:
        return date(year, mon_num, start_day), date(year, mon_num, end_day)
    except ValueError:
        return None


def prorate_week(week_start: date, hours: float,
                 periods: list[tuple[str, date, date]]) -> dict[str, float]:
    """
    Split `hours` across the periods the week overlaps, weighted by
    business-day count in each period.

    Returns {period_label: prorated_hours}
    """
    # Count business days per overlapping period
    bdays = {}
    for label, p_start, p_end in periods:
        bd = _business_days_overlap(week_start, p_start, p_end)
        if bd > 0:
            bdays[label] = bd

    total_bdays = sum(bdays.values())
    if total_bdays == 0:
        return {}

    return {label: round(hours * bd / total_bdays, 4)
            for label, bd in bdays.items()}


# ================================================================
# OPENAIR PARSER
# ================================================================

def _leading_spaces(s: str) -> int:
    return len(s) - len(s.lstrip(' '))


def parse_openair_report(file_obj, default_year: int | None = None) -> dict:
    """
    Parse an OpenAir Excel report.

    Returns nested dict:
        {
            "Brooks": {                              ← last name (first token)
                "Aramark 1099 Reporting Services": { ← project name (stripped)
                    "Apr 1-15":  8.0,
                    "Apr 16-30": 16.0,
                }
            }
        }

    file_obj: file path string OR file-like object (Streamlit UploadedFile)
    """
    import openpyxl
    import calendar

    if default_year is None:
        default_year = date.today().year

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # ---- Parse header row ----
    header = rows[0]
    week_cols = {}   # col_index → week_start date
    for col_idx, val in enumerate(header):
        if val and 'week' in str(val).lower():
            d = _parse_week_date(str(val), default_year)
            if d:
                week_cols[col_idx] = d
        # skip "Total" column automatically

    if not week_cols:
        raise ValueError("No 'Week starting ...' columns found in OpenAir report header.")

    # Determine all months touched by this report → build period bounds per month
    months_in_report = set(d.month for d in week_cols.values())
    period_bounds = {}  # month_num → [(label, start, end), ...]
    for mon in months_in_report:
        last_day = calendar.monthrange(default_year, mon)[1]
        mon_abbr = date(default_year, mon, 1).strftime('%b')
        period_bounds[mon] = [
            (f"{mon_abbr} 1-15",
             date(default_year, mon, 1),
             date(default_year, mon, 15)),
            (f"{mon_abbr} 16-{last_day}",
             date(default_year, mon, 16),
             date(default_year, mon, last_day)),
        ]

    # ---- Parse data rows ----
    result: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    current_person  = None
    current_project = None

    for row in rows[1:]:
        first_cell = row[0]
        if first_cell is None:
            continue

        cell_str   = str(first_cell)
        spaces     = _leading_spaces(cell_str)
        label      = cell_str.strip()

        if not label:
            continue

        if spaces == 0:
            # Person row — extract last name (first token)
            current_person  = label.split()[0]
            current_project = None

        elif spaces <= 2:
            # Client row — aggregates only, skip (we track at project level)
            current_project = None

        else:
            # Project row (4+ spaces)
            if current_person is None:
                continue
            current_project = label

            # Sum prorated hours across all week columns
            for col_idx, week_start in week_cols.items():
                raw_hours = row[col_idx]
                try:
                    hours = float(raw_hours) if raw_hours is not None else 0.0
                except (ValueError, TypeError):
                    hours = 0.0

                if hours == 0:
                    continue

                # Determine which periods this week overlaps
                periods_for_week = period_bounds.get(week_start.month, [])
                # Also check adjacent month if week straddles month boundary
                next_mon = week_start.month % 12 + 1
                next_year = default_year + (1 if next_mon == 1 else 0)
                if next_mon in period_bounds:
                    periods_for_week = periods_for_week + period_bounds[next_mon]

                prorated = prorate_week(week_start, hours, periods_for_week)
                for period_label, ph in prorated.items():
                    result[current_person][current_project][period_label] += ph

    # Convert defaultdicts to plain dicts
    return {
        person: {
            proj: dict(periods)
            for proj, periods in projects.items()
        }
        for person, projects in result.items()
    }


# ================================================================
# SCHEDULE READER
# ================================================================

def read_schedule_hours(wb, sheet_name: str) -> dict:
    """
    Read scheduled hours from a month tab.

    Returns same nested structure as parse_openair_report:
        {person_last_name: {project_code: {period_label: hours}}}
    """
    ws = wb[sheet_name]

    PERSON_NAME_ROW  = 2
    PERIOD_LABEL_ROW = 6
    DATA_START_ROW   = 8
    PERSON_COL_START = 7
    PERSON_COL_END   = 32

    col_map = {}
    for col in range(PERSON_COL_START, PERSON_COL_END + 1):
        period_val = ws.cell(row=PERIOD_LABEL_ROW, column=col).value
        period_str = str(period_val).strip() if period_val else ""
        pair_start = PERSON_COL_START + ((col - PERSON_COL_START) // 2) * 2
        person_val = ws.cell(row=PERSON_NAME_ROW, column=pair_start).value
        person_str = str(person_val).strip() if person_val else ""
        if person_str and period_str:
            col_map[col] = (person_str, period_str)

    schedule = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    consecutive_blank = 0

    for row_num in range(DATA_START_ROW, 5000):
        project_code = ws.cell(row=row_num, column=2).value
        client       = ws.cell(row=row_num, column=1).value
        if not client and not project_code:
            consecutive_blank += 1
            if consecutive_blank >= 10:
                break
            continue
        consecutive_blank = 0

        code_str = str(project_code).strip() if project_code else ""
        for col, (person, period) in col_map.items():
            val = ws.cell(row=row_num, column=col).value
            try:
                hours = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                hours = 0.0
            if hours != 0:
                schedule[person][code_str][period] += hours

    return {p: {proj: dict(periods) for proj, periods in projs.items()}
            for p, projs in schedule.items()}


# ================================================================
# PROJECT CODE MATCHING
# ================================================================

def _normalize(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _match_project(openair_name: str, schedule_codes: list[str]) -> str | None:
    """
    Try to match an OpenAir project name to a schedule project code.
    1. Exact match (case-insensitive, stripped)
    2. Normalized match (remove punctuation/spaces)
    3. One contains the other (longest overlap wins)
    Returns the best matching schedule code, or None.
    """
    oa_norm = _normalize(openair_name)
    best    = None
    best_len = 0

    for sc in schedule_codes:
        sc_norm = _normalize(sc)
        if oa_norm == sc_norm:
            return sc                          # exact — return immediately
        # Containment
        if oa_norm in sc_norm or sc_norm in oa_norm:
            overlap = min(len(oa_norm), len(sc_norm))
            if overlap > best_len:
                best, best_len = sc, overlap

    return best


# ================================================================
# VARIANCE CALCULATOR
# ================================================================

def compute_variances(actual_data: dict, schedule_data: dict,
                      threshold: float = VARIANCE_THRESHOLD) -> list:
    """
    Compare per-person / per-project / per-period actuals vs schedule.
    Returns list of flagged variance dicts (|diff| > threshold).

    Handles fuzzy project code matching between OpenAir and schedule.
    """
    variances = []

    for person, actual_projects in actual_data.items():
        # Find matching person in schedule (exact first, then case-insensitive)
        sched_person = None
        for sp in schedule_data:
            if sp.lower() == person.lower():
                sched_person = sp
                break
        if sched_person is None:
            continue

        sched_projects = schedule_data[sched_person]
        sched_code_list = list(sched_projects.keys())

        for oa_project, actual_periods in actual_projects.items():
            # Match OpenAir project name → schedule project code
            matched_code = _match_project(oa_project, sched_code_list)
            sched_periods = sched_projects.get(matched_code, {}) if matched_code else {}

            all_periods = set(actual_periods) | set(sched_periods)
            for period in all_periods:
                actual_hrs = round(actual_periods.get(period, 0.0), 1)
                sched_hrs  = round(sched_periods.get(period, 0.0), 1)
                diff       = round(actual_hrs - sched_hrs, 1)

                if abs(diff) > threshold:
                    question = (
                        "Do additional hours need to be added to the "
                        "schedule for this project?"
                        if diff < 0 else
                        "Are the scheduled hours expected to hit? "
                        "Or do the hours need to be pushed to another period?"
                    )
                    variances.append({
                        "person":       person,
                        "project_code": matched_code or oa_project,
                        "oa_name":      oa_project,
                        "period":       period,
                        "actual_hours": actual_hrs,
                        "sched_hours":  sched_hrs,
                        "difference":   diff,
                        "question":     question,
                    })

    return variances
