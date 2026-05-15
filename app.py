from __future__ import annotations
# ============================================================
# app.py — GTM Scheduling Analyzer  |  streamlit run app.py
# ============================================================

import gc
import hashlib
import io
import warnings
from collections import defaultdict
from datetime import datetime

import openpyxl
import streamlit as st
import streamlit.components.v1 as components

from config import (
    SENDER_EMAIL, SENDER_NAMES, EMAIL_LOOKUP, INTERN_NAMES, STAFF_NAMES, NAME_ALIASES,
    POSITION_ORDER, PERSON_ROLE, _rank, DISPLAY_NAMES,
    DEFAULT_BUDGET_THRESHOLD, DEFAULT_NEGATIVE_THRESHOLD,
    DEFAULT_PROJECTION_THRESHOLD_PCT,
    DEFAULT_VARIANCE_MIN, DEFAULT_VARIANCE_MAX,
)
from email_utils import (
    email_configured, EMAIL_OK, send_emails_batch, build_html_email,
)
from processors.budget_actual   import process_budget_actual
from processors.project_tracker import process_project_tracker
from processors.variance        import (
    parse_openair_report, read_schedule_hours,
    compute_variances, get_available_months, filter_by_months,
    get_schedule_periods,
)
from processors.utilization import process_utilization, get_pto_schedule

warnings.filterwarnings("ignore", category=UserWarning)

st.set_page_config(page_title="GTM Scheduling Analyzer",
                   layout="wide", page_icon="📊")

st.markdown("""<style>
[data-testid="metric-container"] {
    background:#f8fafc; border:1px solid #e2e8f0;
    border-radius:10px; padding:16px 20px;
    box-shadow:0 1px 3px rgba(0,0,0,.06);
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    font-size:13px; color:#64748b; font-weight:500;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size:28px; font-weight:700; color:#0E2841;
}
.stTabs [data-baseweb="tab-list"] {
    gap:4px; background:#f1f5f9; border-radius:10px; padding:4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius:8px; padding:8px 20px; font-weight:500; color:#64748b;
}
.stTabs [aria-selected="true"] {
    background:white !important; color:#0E2841 !important;
    box-shadow:0 1px 3px rgba(0,0,0,.1);
}
[data-testid="stSidebar"] { background:#f8fafc; border-right:1px solid #e2e8f0; }
[data-testid="stExpander"] { border:1px solid #e2e8f0 !important; border-radius:8px !important; }
</style>""", unsafe_allow_html=True)

import os as _os
_logo_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "assets", "logo.png")
if _os.path.exists(_logo_path):
    try:
        st.logo(_logo_path)
    except Exception:
        pass

# Header with logo embedded via HTML to prevent column clipping
_logo_b64 = ""
if _os.path.exists(_logo_path):
    import base64 as _b64
    with open(_logo_path, "rb") as _f:
        _logo_b64 = _b64.b64encode(_f.read()).decode()

if _logo_b64:
    st.markdown(
        f'''<div style="display:flex;align-items:center;gap:16px;margin-bottom:4px;">
        <img src="data:image/png;base64,{_logo_b64}"
             style="height:72px;width:auto;object-fit:contain;flex-shrink:0;">
        <div>
          <div style="font-size:1.8rem;font-weight:700;color:#0E2841;line-height:1.2;">GTM Scheduling Analyzer</div>
          <div style=\"font-size:0.85rem;color:#64748b;\">Today: {datetime.now().strftime('%A, %B %d, %Y')}</div>
        </div></div>''',
        unsafe_allow_html=True)
else:
    st.markdown("## GTM Scheduling Analyzer")
    st.caption(f"Today: {datetime.now().strftime('%A, %B %d, %Y')}")

# ============================================================
# AUTHENTICATION
# Credentials stored in Streamlit Secrets:
#   [auth]
#   username = "gtmtas"
#   password_hash = "77c0b600dc99b2c0b5dc5db009c929f16927148a53cd11b4be986237599f69ee"
# Falls back to hardcoded hash if secrets not configured.
# ============================================================
import hashlib as _hl

_FALLBACK_USER = "gtmtas"
_FALLBACK_HASH = "77c0b600dc99b2c0b5dc5db009c929f16927148a53cd11b4be986237599f69ee"

def _check_credentials(username: str, password: str) -> bool:
    pw_hash = _hl.sha256(password.encode()).hexdigest()
    try:
        stored_user = st.secrets["auth"]["username"]
        stored_hash = st.secrets["auth"]["password_hash"]
    except (KeyError, FileNotFoundError):
        stored_user = _FALLBACK_USER
        stored_hash = _FALLBACK_HASH
    return username.strip() == stored_user and pw_hash == stored_hash

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown("<div style='height:40px'></div>", unsafe_allow_html=True)
    _, login_col, _ = st.columns([1, 1.2, 1])
    with login_col:
        if _logo_b64:
            st.markdown(
                f'<div style="text-align:center;margin-bottom:8px;">'
                f'<img src="data:image/png;base64,{_logo_b64}" style="height:80px;width:auto;"></div>',
                unsafe_allow_html=True)
        st.markdown(
            '<div style="text-align:center;font-size:1.4rem;font-weight:700;'
            'color:#0E2841;margin-bottom:4px;">GTM Scheduling Analyzer</div>',
            unsafe_allow_html=True)
        st.markdown(
            '<div style="text-align:center;color:#64748b;margin-bottom:24px;'
            'font-size:0.9rem;">Sign in to continue</div>',
            unsafe_allow_html=True)
        with st.form("login_form"):
            username = st.text_input("Username", placeholder="Enter username")
            password = st.text_input("Password", type="password", placeholder="Enter password")
            submitted = st.form_submit_button("Sign In", use_container_width=True, type="primary")
            if submitted:
                if _check_credentials(username, password):
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("Incorrect username or password.")
    st.stop()

# ── Logout button in sidebar ──────────────────────────────────
# (rendered after authentication check so sidebar only shows when logged in)

if not email_configured():
    st.warning("⚠️ **Email credentials not configured.** "
               "Add SendGrid keys in Streamlit Secrets to enable sending.")

# ============================================================
# HELPERS
# ============================================================

def _hash(b: bytes) -> str:
    return hashlib.md5(b).hexdigest()

def _sender_name() -> str:
    try:
        from_email = st.secrets["email"]["from_email"]
    except (KeyError, FileNotFoundError):
        from_email = SENDER_EMAIL
    return SENDER_NAMES.get(from_email, "Jake")

def _find_sheet(sheetnames, keywords):
    for name in sheetnames:
        if any(kw in name.lower() for kw in keywords):
            return name
    return None

def _lookup_email(name: str):
    if not name:
        return None
    name = name.strip()
    if name in EMAIL_LOOKUP:
        return EMAIL_LOOKUP[name]
    for k, v in EMAIL_LOOKUP.items():
        if k.lower() == name.lower():
            return v
    return None

def _normalize_name(name: str) -> str:
    """Resolve name aliases so bare 'O\'Donnell' → 'J. O\'Donnell'."""
    if not name:
        return name
    return NAME_ALIASES.get(name.strip(), name.strip())

# ============================================================
# SESSION STATE
# ============================================================
_DEFAULTS = {
    "run_triggered":   False,
    "sched_bytes":     None,
    "oa_bytes":        None,
    "selected_owners": set(),
    "settings": {
        "budget_threshold":   float(DEFAULT_BUDGET_THRESHOLD),
        "negative_threshold": float(DEFAULT_NEGATIVE_THRESHOLD),
        "variance_min":       float(DEFAULT_VARIANCE_MIN),
        "variance_max":       float(DEFAULT_VARIANCE_MAX),
    },
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ============================================================
# CACHING
# cache_resource — for complex objects (workbooks, nested dicts)
#                  stored in memory, NOT pickled → no crash
# cache_data     — for simple serialisable returns (lists, tuples of strings)
# ============================================================

@st.cache_resource(max_entries=2, show_spinner=False)
def _load_wb(file_hash: str, _file_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(_file_bytes), data_only=True)

@st.cache_resource(max_entries=2, show_spinner=False)
def _parse_openair(oa_hash: str, _oa_bytes: bytes) -> dict:
    """Parse OpenAir CSV and cache the nested dict as a resource (no pickling)."""
    return parse_openair_report(io.BytesIO(_oa_bytes))

@st.cache_data(show_spinner=False)
def run_budget(file_hash, _b, budget_thr, proj_pct, neg_thr, _v="v2"):  # bump to bust cache after logic changes
    wb    = _load_wb(file_hash, _b)
    sheet = _find_sheet(wb.sheetnames, ["budget to actual", "budget"])
    if not sheet:
        return [], None
    result = process_budget_actual(wb[sheet], budget_thr, proj_pct, neg_thr)
    gc.collect()
    return result, sheet

@st.cache_data(show_spinner=False)
def run_tracker(file_hash, _b):
    wb    = _load_wb(file_hash, _b)
    sheet = _find_sheet(wb.sheetnames, ["project tracker", "tracker"])
    if not sheet:
        return [], [], None
    issues, tbd = process_project_tracker(wb[sheet])
    gc.collect()
    return issues, tbd, sheet

@st.cache_data(show_spinner=False)
def run_utilization(file_hash, _b, month):
    wb = _load_wb(file_hash, _b)
    try:
        data = process_utilization(wb, target_month=month)
    except Exception:
        data = []
    gc.collect()
    return data

@st.cache_data(show_spinner=False)
def get_valid_people(file_hash, _b, active_month):
    wb = _load_wb(file_hash, _b)
    for name in wb.sheetnames:
        if name.lower().startswith(active_month[:3].lower()):
            ws = wb[name]
            people = set()
            for col in range(7, 45):
                val = ws.cell(row=2, column=col).value
                if val:
                    people.add(_normalize_name(str(val).strip()))
            return people
    return set()

def get_oa_periods(oa_hash, _oa_bytes):
    """Return (available, future) period lists from OpenAir data.
    Not cached — period logic is fast; underlying parse IS cached in _parse_openair."""
    actual = _parse_openair(oa_hash, _oa_bytes)
    return get_available_months(actual)

def get_sched_periods(file_hash, _b, active_month):
    """Return (available, future) period lists from the schedule sheet.
    Not cached — period logic is fast; underlying workbook IS cached in _load_wb."""
    wb = _load_wb(file_hash, _b)
    sheet = next(
        (s for s in wb.sheetnames if s.lower().startswith(active_month[:3].lower())),
        None,
    )
    if not sheet:
        return [], []
    return get_schedule_periods(wb, sheet)

@st.cache_data(show_spinner=False)
def run_variance(file_hash, _b, oa_hash, _oa,
                 selected_months_tuple, var_min, var_max, active_month,
                 _v="v13"):  # bump _v to bust stale cache after code changes
    wb = _load_wb(file_hash, _b)
    sched_sheet = next(
        (s for s in wb.sheetnames if s.lower().startswith(active_month[:3].lower())),
        None,
    )
    if not sched_sheet:
        return [], f"No sheet found for {active_month}"
    try:
        sched = read_schedule_hours(wb, sched_sheet)
        if oa_hash and _oa:
            # Use real OpenAir actuals
            actual   = _parse_openair(oa_hash, _oa)
            filtered = filter_by_months(actual, list(selected_months_tuple))
        else:
            # No OpenAir — build fake actual_data with 0s from the schedule
            # so that any scheduled hours will produce a variance (actual=0).
            actual = {}
            for person, projects in sched.items():
                actual[person] = {}
                for proj, periods in projects.items():
                    actual[person][proj] = {period: 0.0 for period in periods}
            filtered = filter_by_months(actual, list(selected_months_tuple))
        variances = compute_variances(filtered, sched,
                                      min_diff=var_min, max_diff=var_max,
                                      selected_periods=list(selected_months_tuple))
        gc.collect()
        return variances, None
    except Exception as e:
        gc.collect()
        return [], str(e)

# ============================================================
# SIDEBAR — settings (st.form prevents reruns on +/- clicks)
# ============================================================
active_month = datetime.now().strftime("%B")
sender_name  = _sender_name()

# Compute current month + next 2 for PTO schedule
def _pto_months(current: str) -> list:
    import calendar
    month_names = list(calendar.month_name)[1:]  # Jan..Dec
    try:
        idx = month_names.index(current)
    except ValueError:
        return [current]
    return [month_names[(idx + i) % 12] for i in range(3)]

_pto_month_list = _pto_months(active_month)

with st.sidebar:
    _lcol1, _lcol2 = st.columns([0.65, 0.35])
    with _lcol1:
        st.header("⚙️ Settings")
    with _lcol2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Sign Out", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()
    with st.form("settings_form"):
        st.subheader("💰 Budget to Actual")
        f_budget = st.number_input(
            "Flag unscheduled remaining over ($)",
            value=int(st.session_state.settings["budget_threshold"]),
            step=1000, min_value=0,
        )
        f_negative = st.number_input(
            "Flag negative budgets below -($)",
            value=int(st.session_state.settings["negative_threshold"]),
            step=50, min_value=0,
        )
        st.divider()
        st.subheader("📊 Variance Thresholds")
        applied = st.form_submit_button(
            "✔ Apply Budget Settings", type="primary", use_container_width=True)

    if applied:
        st.session_state.settings["budget_threshold"]   = float(f_budget)
        st.session_state.settings["negative_threshold"] = float(f_negative)
        st.rerun()

    # ── Variance thresholds — outside form for real-time slider↔input sync
    st.divider()
    st.subheader("📊 Variance Thresholds")

    if "_vmin" not in st.session_state:
        st.session_state._vmin = float(st.session_state.settings["variance_min"])
    if "_vmax" not in st.session_state:
        st.session_state._vmax = float(st.session_state.settings["variance_max"])

    # on_change callbacks keep slider ↔ number input in sync
    def _sl_min_changed():
        st.session_state._vmin = st.session_state._sl_min
        st.session_state._ni_min = st.session_state._sl_min
    def _ni_min_changed():
        v = float(st.session_state._ni_min or 0.0)
        st.session_state._vmin = v
        st.session_state._sl_min = min(v, 50.0)
    def _sl_max_changed():
        st.session_state._vmax = st.session_state._sl_max
        st.session_state._ni_max = st.session_state._sl_max
    def _ni_max_changed():
        v = float(st.session_state._ni_max or 0.0)
        st.session_state._vmax = v
        st.session_state._sl_max = min(v, 50.0)

    st.markdown("**📉 Scheduled but not actual**")
    st.caption("Flag when scheduled hrs exceed actual hrs by more than:")
    st.slider("", min_value=0.0, max_value=50.0, step=0.5, format="%.1f hrs",
              value=min(st.session_state._vmin, 50.0), key="_sl_min",
              label_visibility="collapsed", on_change=_sl_min_changed)
    st.number_input("Exact (hrs):", min_value=0.0, step=0.5,
                    value=float(st.session_state._vmin or 0.0),
                    key="_ni_min", on_change=_ni_min_changed)

    st.markdown("**📈 Actual but not scheduled**")
    st.caption("Flag when actual hrs exceed scheduled hrs by more than:")
    st.slider("", min_value=0.0, max_value=50.0, step=0.5, format="%.1f hrs",
              value=min(st.session_state._vmax, 50.0), key="_sl_max",
              label_visibility="collapsed", on_change=_sl_max_changed)
    st.number_input("Exact (hrs):", min_value=0.0, step=0.5,
                    value=float(st.session_state._vmax or 0.0),
                    key="_ni_max", on_change=_ni_max_changed)

    if st.button("✔ Apply Variance Settings", type="primary", use_container_width=True):
        st.session_state.settings["variance_min"] = st.session_state._vmin
        st.session_state.settings["variance_max"] = st.session_state._vmax
        st.rerun()

    st.divider()
    st.subheader("📋 Email Lookup")
    st.caption(f"{len(EMAIL_LOOKUP)} people configured")
    if st.checkbox("Show lookup table"):
        st.dataframe(
            [{"Name": k, "Email": v} for k, v in EMAIL_LOOKUP.items()],
            use_container_width=True, hide_index=True,
        )

if not EMAIL_OK:
    st.sidebar.warning("SendGrid keys not found — previews work, sending disabled.")

budget_threshold   = st.session_state.settings["budget_threshold"]
negative_threshold = st.session_state.settings["negative_threshold"]
variance_min       = st.session_state.settings["variance_min"]
variance_max       = st.session_state.settings["variance_max"]

# ============================================================
# FILE UPLOAD + RUN BUTTON
#
# Before run:   show uploaders + Run button
# After run:    hide uploaders, lock bytes in session state
#               show "Change Files" button to reset
# ============================================================
st.subheader("📁 Upload Files")

if not st.session_state.run_triggered:
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        schedule_file = st.file_uploader(
            "Schedule File (.xlsx) — required",
            type=["xlsx", "csv"], key="schedule")
    with col_f2:
        openair_file = st.file_uploader(
            "OpenAir Report (.csv or .xlsx) — optional",
            type=["csv", "xlsx"], key="openair")

    run_col, _ = st.columns([0.25, 0.75])
    with run_col:
        run_clicked = st.button(
            "▶ Run Analysis",
            type="primary",
            disabled=not schedule_file,
            use_container_width=True,
        )

    if run_clicked and schedule_file:
        with st.spinner("Reading files…"):
            try:
                st.session_state.sched_bytes = bytes(schedule_file.read())
                st.session_state.oa_bytes = (
                    bytes(openair_file.read()) if openair_file else None
                )
                st.session_state.run_triggered = True
                st.session_state._analysis_done = False  # show spinner on fresh run
            except Exception as e:
                st.error(f"Could not read files: {e}")
        st.rerun()
    else:
        st.info("Upload your schedule file then click **▶ Run Analysis** to begin.")
        st.stop()

else:
    # Show change-files button but NOT while analysis is running
    chg_col, _ = st.columns([0.4, 0.6])
    with chg_col:
        if st.button("🔄 Change Files / Run Again", use_container_width=True):
            st.session_state.run_triggered = False
            st.session_state.sched_bytes   = None
            st.session_state.oa_bytes      = None
            st.rerun()

sched_bytes = st.session_state.sched_bytes
oa_bytes    = st.session_state.oa_bytes
file_hash   = _hash(sched_bytes)
has_openair = bool(oa_bytes)
oa_hash     = _hash(oa_bytes) if oa_bytes else ""

# ============================================================
# ANALYSIS — all runs under a single spinner to block UI
# ============================================================
_show_spinner = not st.session_state.get("_analysis_done", False)
with st.spinner("🔄 Running analysis — please wait…") if _show_spinner else st.empty():

    try:
        _wb_check = _load_wb(file_hash, sched_bytes)
        sheets    = list(_wb_check.sheetnames)
    except Exception as e:
        st.error(f"Could not open schedule file: {e}")
        st.stop()

    budget_issues, budget_sheet_name = [], None
    try:
        budget_issues, budget_sheet_name = run_budget(
            file_hash, sched_bytes,
            budget_threshold, DEFAULT_PROJECTION_THRESHOLD_PCT, negative_threshold,
        )
    except Exception as e:
        st.warning(f"Budget error: {e}")

    tracker_issues, tbd_projects, tracker_sheet_name = [], [], None
    try:
        tracker_issues, tbd_projects, tracker_sheet_name = run_tracker(
            file_hash, sched_bytes)
    except Exception as e:
        st.warning(f"Tracker error: {e}")

    util_data = []
    try:
        util_data = run_utilization(file_hash, sched_bytes, active_month)
    except Exception as e:
        st.warning(f"Utilization error: {e}")

    pto_schedule_data = {}
    try:
        _wb_pto = _load_wb(file_hash, sched_bytes)
        pto_schedule_data = get_pto_schedule(_wb_pto, _pto_month_list)
    except Exception:
        pass

    valid_people = set()
    try:
        valid_people = get_valid_people(file_hash, sched_bytes, active_month)
    except Exception:
        pass
    st.session_state._analysis_done = True  # suppress spinner on settings reruns

    # OpenAir or schedule-derived periods
    available_months, future_months = [], []
    openair_error = None
    try:
        if has_openair:
            available_months, future_months = get_oa_periods(oa_hash, oa_bytes)
        else:
            available_months, future_months = get_sched_periods(
                file_hash, sched_bytes, active_month)
    except Exception as e:
        openair_error = str(e)

st.success(f"✅ Loaded **{len(sheets)}** sheet(s)")
c1, c2, c3 = st.columns(3)
c1.info(f"💰 Budget: **{budget_sheet_name or 'Not found'}**")
c2.info(f"📋 Tracker: **{tracker_sheet_name or 'Not found'}**")
c3.info(f"📅 Month: **{active_month}**")

if not has_openair:
    st.info("ℹ️ No OpenAir report uploaded — variance will show scheduled hours "
            "with actual hours as 0. Upload an OpenAir report for real actuals.")

st.divider()

# ============================================================
# PERIOD SELECTOR + VARIANCE
# ============================================================
selected_months = []
variance_issues = []
var_error       = None

if available_months:
    current_abbr   = datetime.now().strftime("%b")
    default_months = [m for m in available_months
                      if m.startswith(current_abbr) and m not in future_months]
    if not default_months:
        default_months = [m for m in available_months if m not in future_months][-1:]

    st.subheader("📅 Variance Period Selection")

    def _period_label(p):
        if p in future_months:
            return f"🔮 {p} (future — no actuals yet)"
        return p

    selected_months = st.multiselect(
        "Select period(s) for variance analysis:",
        options=available_months,
        default=default_months,
        format_func=_period_label,
        help=(
            "All 24 half-month periods for the year are shown. "
            "🔮 = future (scheduled hours only, actual = 0). "
            "Past periods use OpenAir actuals if uploaded."
        ),
    )

    if selected_months:
        with st.spinner("Computing variances…"):
            try:
                variance_issues, var_error = run_variance(
                    file_hash, sched_bytes,
                    oa_hash, oa_bytes,
                    tuple(selected_months),
                    -variance_min, variance_max,  # min is stored positive, negated here
                    active_month,
                )
            except Exception as e:
                var_error = str(e)

# ============================================================
# OWNER MAP
# Build owners_data for everyone in valid_people who has an email.
# This ensures ALL schedule staff appear in the email list.
# ============================================================
owners_data = defaultdict(lambda: {
    "email": None, "first_name": "there",
    "tracker": [], "budget": [], "variance": [], "util": [],
})

# Seed ALL valid_people who have emails so nobody is missed
for person in valid_people:
    email = _lookup_email(person)
    if email and not owners_data[person]["email"]:
        from config import FIRST_NAMES
        owners_data[person]["email"]      = email
        owners_data[person]["first_name"] = FIRST_NAMES.get(person, FIRST_NAMES.get(NAME_ALIASES.get(person, ""), person))

for issue in tracker_issues:
    o = _normalize_name(issue.get("owner", ""))
    if not o or (valid_people and o not in valid_people):
        continue
    if not owners_data[o]["email"]:
        owners_data[o]["email"]      = issue.get("owner_email")
        owners_data[o]["first_name"] = issue.get("owner_first", o)
    owners_data[o]["tracker"].append(issue)

for issue in budget_issues:
    o = _normalize_name(issue.get("owner", ""))
    if not o or (valid_people and o not in valid_people):
        continue
    if not owners_data[o]["email"]:
        owners_data[o]["email"]      = issue.get("owner_email")
        owners_data[o]["first_name"] = issue.get("owner_first", o)
    owners_data[o]["budget"].append(issue)

# Build project → owner map from tracker + budget data
_project_owner_map = {}
for _issue in tracker_issues:
    _code  = str(_issue.get("project_code", "")).strip()
    _owner = _normalize_name(_issue.get("owner", ""))
    if _code and _owner:
        _project_owner_map[_code] = _owner
for _issue in budget_issues:
    _code  = str(_issue.get("project_code", "")).strip()
    _owner = _normalize_name(_issue.get("owner", ""))
    if _code and _owner:
        _project_owner_map.setdefault(_code, _owner)

for v in variance_issues:
    person       = _normalize_name(v.get("person", ""))
    project_code = v.get("project_code", "")
    proj_owner   = _normalize_name(_project_owner_map.get(project_code, ""))

    # Staff members get their OWN variance rows in their personal email
    if person in STAFF_NAMES and person and (not valid_people or person in valid_people):
        if not owners_data[person]["email"]:
            owners_data[person]["email"] = _lookup_email(person)
        owners_data[person]["variance"].append(v)

    # Project owner gets ALL rows for their projects (including staff rows)
    if proj_owner and proj_owner not in STAFF_NAMES and (not valid_people or proj_owner in valid_people):
        if not owners_data[proj_owner]["email"]:
            owners_data[proj_owner]["email"] = _lookup_email(proj_owner)
        owners_data[proj_owner]["variance"].append(v)
    elif not proj_owner and person not in STAFF_NAMES:
        # No project owner found — fallback to person themselves
        if person and (not valid_people or person in valid_people):
            if not owners_data[person]["email"]:
                owners_data[person]["email"] = _lookup_email(person)
            owners_data[person]["variance"].append(v)

for u in util_data:
    p = _normalize_name(u.get("person", ""))
    if not p or (valid_people and p not in valid_people):
        continue
    if not owners_data[p]["email"]:
        owners_data[p]["email"]      = u.get("person_email") or _lookup_email(p)
        owners_data[p]["first_name"] = u.get("first_name", p)
    owners_data[p]["util"].append(u)

# active_owners = everyone in valid_people who has an email
active_owners = {
    owner: data for owner, data in owners_data.items()
    if data.get("email")
    and (not valid_people or owner in valid_people)
}

# ============================================================
# ANALYSIS TABS
# ============================================================
tab1, tab2, tab3, tab4 = st.tabs([
    "📋 Project Tracker", "💰 Budget to Actual",
    "📈 Utilization",     "📊 Variance (OpenAir)"])

with tab1:
    st.header("Project Tracker — Known Projects")
    if not tracker_sheet_name:
        st.error("No Project Tracker tab found.")
    else:
        c1, c2 = st.columns(2)
        c1.metric("⚠️ Issues",           len(tracker_issues))
        c2.metric("📌 TBD / Pending SOW", len(tbd_projects))
        if tbd_projects:
            with st.expander(f"📌 {len(tbd_projects)} TBD / Pending SOW projects"):
                st.dataframe(
                    [{"Client": p.get("client",""), "Project Code": p.get("project_code",""),
                      "Status": p.get("status","TBD"), "Owner": p.get("owner",""),
                      "Budget": f"${p.get('budget',0):,.0f}"}
                     for p in tbd_projects],
                    use_container_width=True, hide_index=True)
        if not tracker_issues:
            st.success("✅ No issues found!")
        else:
            rows = []
            for i in tracker_issues:
                missing = i.get("missing_rates", [])
                for m in missing:
                    rows.append({"Client": i.get("client",""),
                                 "Project Code": i.get("project_code",""),
                                 "Owner": i.get("owner",""),
                                 "To Be Reviewed": f"Missing {m} Rate",
                                 "Has Email": "✅" if i.get("owner_email") else "❌"})
                if not missing:
                    for prob in i.get("problems", []):
                        rows.append({"Client": i.get("client",""),
                                     "Project Code": i.get("project_code",""),
                                     "Owner": i.get("owner",""),
                                     "To Be Reviewed": prob,
                                     "Has Email": "✅" if i.get("owner_email") else "❌"})
            st.dataframe(rows, use_container_width=True, hide_index=True)

with tab2:
    st.header("Budget to Actual — Known Projects")
    st.caption(f"Flagging: negative ≤ -${negative_threshold:,.0f} | "
               f"unscheduled > ${budget_threshold:,.0f}")
    if not budget_sheet_name:
        st.error("No Budget to Actual tab found.")
    else:
        neg = [i for i in budget_issues if i.get("type") == "negative"]
        np_ = [i for i in budget_issues if i.get("type") == "not_projected"]
        c1, c2, c3 = st.columns(3)
        c1.metric("🔴 Over Budget", len(neg))
        c2.metric("🟡 Unscheduled", len(np_))
        c3.metric("⚠️ Total",       len(budget_issues))
        if budget_issues:
            st.dataframe(
                [{"Client": i.get("client",""), "Project Code": i.get("project_code",""),
                  "Owner": i.get("owner",""),
                  "Budget": f"${i.get('budget',0):,.0f}",
                  "Remaining": f"${i.get('remaining',0):,.0f}",
                  "Flag": i.get("description",""),
                  "Has Email": "✅" if i.get("owner_email") else "❌"}
                 for i in budget_issues],
                use_container_width=True, hide_index=True)

with tab3:
    st.header(f"Utilization — {active_month}")
    if not util_data:
        st.warning("No utilization data found. Make sure the workbook has a "
                   "'Utilization by Month' tab.")
    else:
        st.dataframe(
            [{"Role": u.get("role",""), "Person": u.get("person",""),
              "Chargeable": u.get("chargeable","-"), "Holiday": u.get("holiday","-"),
              "PTO": u.get("pto","-"), "Month Total": u.get("month_total","-"),
              "Remaining": u.get("remaining","-"),
              "Utilization": f"{u['utilization_pct']:.1f}%" if u.get("utilization_pct") is not None else "-",
              "Goal":        f"{u['goal_pct']:.0f}%"        if u.get("goal_pct")        is not None else "-",
              "Difference":  f"{u['difference_pct']:+.1f}%" if u.get("difference_pct") is not None else "-",
              "Has Email":   "✅" if u.get("person_email") else "❌"}
             for u in util_data],
            use_container_width=True, hide_index=True)

with tab4:
    st.header("Actual vs Schedule Variance (OpenAir)")
    if not has_openair:
        st.info("ℹ️ No OpenAir report uploaded — showing scheduled hours with actual = 0.")
    if openair_error:
        st.error(f"Error loading periods: {openair_error}")
    if var_error:
        st.error(f"Variance error: {var_error}")
    if not selected_months:
        st.warning("Select at least one period above.")
    elif not variance_issues:
        st.success(f"✅ No variances outside "
                   f"[{variance_min:+.0f}, {variance_max:+.0f}] hrs.")
    else:
        if len(selected_months) > 1:
            st.info(f"Showing across {len(selected_months)} periods: "
                    f"{', '.join(selected_months)}")
        st.metric("⚠️ Variances Found", len(variance_issues))
        _sorted_v = sorted(variance_issues, key=lambda v: (_rank(v.get("person","")), v.get("project_code","")))
        st.dataframe(
            [{"Person": v.get("person",""), "Project": v.get("project_code",""),
              "Period": v.get("period",""),
              "Actual Hrs": v.get("actual_hours",0), "Sched Hrs": v.get("sched_hours",0),
              "Diff": v.get("difference",0), "To Review": v.get("question",""),
              "Future": "🔮" if v.get("is_future") else ""}
             for v in _sorted_v],
            use_container_width=True, hide_index=True)

# ============================================================
# COMBINED EMAILS
# ============================================================
st.divider()
st.header("📧 Combined Emails")
st.caption("One email per person · Project Tracker · Budget · "
           "TBD/Pending SOW · Variance · Utilization")

if not active_owners:
    st.info("No staff found with email addresses — check config.py EMAIL_LOOKUP.")
    st.stop()

st.metric("TAS Members", len(active_owners))
all_owner_keys = sorted(active_owners.keys(), key=_rank)

if not st.session_state.selected_owners.issubset(set(all_owner_keys)):
    st.session_state.selected_owners = set(all_owner_keys)

col_sa, col_da, _ = st.columns([0.15, 0.18, 0.67])
with col_sa:
    if st.button("✅ Select All"):
        st.session_state.selected_owners = set(all_owner_keys)
        st.rerun()
with col_da:
    if st.button("⬜ Deselect All"):
        st.session_state.selected_owners = set()
        st.rerun()

st.markdown("**Select recipients:**")
for owner in all_owner_keys:
    data       = active_owners[owner]
    first_name = data.get("first_name", owner)
    email_str  = data.get("email") or "⚠️ no email"
    _display = DISPLAY_NAMES.get(owner, owner)
    label = (f"**{first_name} ({_display})** · {email_str} — "
             f"Tracker: {len(data['tracker'])} · "
             f"Budget: {len(data['budget'])} · "
             f"Util: {len(data['util'])} · "
             f"Variance: {len(data['variance'])}")
    checked = owner in st.session_state.selected_owners
    if st.checkbox(label, value=checked, key=f"chk_{owner}"):
        st.session_state.selected_owners.add(owner)
    else:
        st.session_state.selected_owners.discard(owner)

# ---- Build HTML emails ----
st.markdown("**Email previews:**")
combined_emails = []

for owner in sorted(st.session_state.selected_owners, key=_rank):
    if owner not in active_owners:
        continue
    data         = active_owners[owner]
    person_email = data.get("email")
    first_name   = data.get("first_name", owner)
    if not person_email:
        continue

    is_intern     = owner in INTERN_NAMES
    tracker_list  = data.get("tracker", [])
    budget_list   = data.get("budget", [])
    variance_list = data.get("variance", [])

    # Interns only see their own variance rows
    if is_intern:
        variance_list = [v for v in variance_list if v.get("person") == owner]

    html = build_html_email(
        owner         = owner,
        first_name    = first_name,
        tracker_issues= tracker_list,
        budget_issues = budget_list,
        tbd_projects  = tbd_projects,
        variance_issues=variance_list,
        util_data     = util_data,
        pto_schedule  = pto_schedule_data,
        pto_months    = _pto_month_list,
        has_openair   = has_openair,
        no_openair_note = not has_openair and bool(variance_list),
        selected_months = selected_months if len(selected_months) > 1 else None,
        is_staff      = owner in STAFF_NAMES,
    )

    if not html:
        continue

    combined_emails.append({
        "to":      person_email,
        "subject": f"Scheduling Review — {active_month}",
        "person":  owner,
        "body":    html,
    })

    with st.expander(f"👁 {first_name} ({owner}) · {person_email}"):
        components.html(html, height=500, scrolling=True)

# ---- Send buttons ----
st.divider()
no_email = [o for o in st.session_state.selected_owners
            if not active_owners.get(o,{}).get("email")]
if no_email:
    names = [active_owners[o].get("first_name",o) for o in no_email
             if o in active_owners]
    st.warning(f"⚠️ No email configured for: {', '.join(names)} — will be skipped.")

sendable = [e for e in combined_emails if e.get("to")]

all_sendable = []
for owner in sorted(all_owner_keys, key=_rank):
    data = active_owners.get(owner, {})
    if not data.get("email"):
        continue
    is_intern     = owner in INTERN_NAMES
    variance_list = data.get("variance", [])
    if is_intern:
        variance_list = [v for v in variance_list if v.get("person") == owner]
    html = build_html_email(
        owner=owner, first_name=data.get("first_name", owner),
        tracker_issues=data.get("tracker",[]),
        budget_issues=data.get("budget",[]),
        tbd_projects=tbd_projects,
        variance_issues=variance_list,
        util_data=util_data,
        pto_schedule=pto_schedule_data,
        pto_months=_pto_month_list,
        has_openair=has_openair,
        no_openair_note=not has_openair and bool(variance_list),
        selected_months=selected_months if len(selected_months) > 1 else None,
        is_staff=owner in STAFF_NAMES,
    )
    if html:
        all_sendable.append({"to": data["email"], "subject": f"Scheduling Review — {active_month}",
                              "person": owner, "body": html})

col_b1, col_b2 = st.columns(2)
with col_b1:
    send_sel = st.button(
        f"📤 Send to Selected ({len(sendable)})",
        type="primary", key="send_selected",
        disabled=not EMAIL_OK or not sendable)
with col_b2:
    send_all_btn = st.button(
        f"📤 Send All ({len(all_sendable)})",
        key="send_all",
        disabled=not EMAIL_OK or not all_sendable)

if not EMAIL_OK:
    st.info("Configure SendGrid keys in Streamlit Secrets to enable sending.")

if send_sel or send_all_btn:
    targets = sendable if send_sel else all_sendable
    with st.spinner(f"Sending {len(targets)} email(s)…"):
        try:
            results = send_emails_batch(targets)
            sent   = [r for r in results if r.get("status") == "sent"]
            failed = [r for r in results if r.get("status") != "sent"]
            if sent:
                st.success(f"✅ {len(sent)} email(s) sent!")
            if failed:
                st.error(f"❌ {len(failed)} failed:")
                for r in failed:
                    st.write(f"  • {r.get('to','?')}: {r.get('status','?')}")
        except Exception as e:
            st.error(f"Error sending: {e}")
